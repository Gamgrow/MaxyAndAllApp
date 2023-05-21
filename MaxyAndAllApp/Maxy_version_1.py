# Maxy


import datetime
import pyttsx3
import smtplib  # for email send
import speech_recognition as sr
import sys
import webbrowser
import os
import subprocess  # for Run any application
import random
import math
import wikipedia
import pywhatkit as pwt  # for youtube video play
import requests  # for temperature get
from googlesearch import search
import openpyxl  # for exel file open and read
import screen_brightness_control as pct  # for brightness

from tkinter import *
wind = Tk()
from PIL import ImageTk, Image
# create date today
date = datetime.datetime.now()
date = str(date)
date = date.split()
date = date[0]

# create random number
r = random.randint(0, 7)


# Create a engine
engine = pyttsx3.init()

# set voice of male or female
voices = engine.getProperty('voices')
engine.setProperty('voice', voices[1].id)

# set voice rate
rate = engine.getProperty('rate')
engine.setProperty('rate', 148)

# create current datetime
cur_time = datetime.datetime.now().strftime('%H:%M:%S')

# take only hour
hour = int(datetime.datetime.now().hour)
minut = int(datetime.datetime.now().minute)

# whether

def shoi():
    pass

def you(str):
  
    mylabel2.config(text=str)

def printtext(str):
    # if var.get() != "":
    #   x = var.get()
    #   mylabel.config(text=x)
    # else:
    
    mylabel.config(text=str)

   
    
    
   


def speak(text):
    engine.say(text)
    engine.runAndWait()


def wishme(text):
    printtext(text)
    printtext()
    if hour >= 0 and hour <= 12:
        printtext('good morning sir,\n I am maxi, \nhow may i help you')
        speak('good morning sir, I am maxi, \n  how may i help you')

    elif hour >= 12 and hour <= 18:
        printtext('good afternoon sir, \nI am maxi,  how may i help you')
        speak('good afternoon sir, I am maxi,  how may i help you')

    else:
        printtext('good evening sir,\n I am maxi,  how may i help you')
        speak('good evening sir, I am maxi,  how may i help you')


def sendmail(email, subject, contant, name):
    from email.message import EmailMessage
    email_id = "Enter Gmail"
    email_pas = "Enter password"

    msg = EmailMessage()
    msg["subject"] = subject
    msg["from"] = "Lalit kumar yadav"
    msg["to"] = email
    msg.set_content(contant)

    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as smtp:
        smtp.login(email_id, email_pas)
        smtp.send_message(msg)
        # printtext("sent mail succesfully...")
        speak("sent mail succesfully")
        printtext("sent mail succesfully...")


def lower(t):
    return t.lower()


#  take command for do anything
def command():
    r = sr.Recognizer()
    with sr.Microphone() as mic:
        # print('listening...')

        r.pause_threshold = 5
        audio = r.listen(mic, timeout=45)

    try:
        # print('recognizing...')
        text = r.recognize_google(audio)
        text = text.lower()
    except Exception as e:
        printtext(e)
        return text
    return text


def doooo(text):
    you(text)
    if text in ["exit", "quit", "stop"]:
                exit()
    elif text in ['what is the time', "what's the time", 'is time', 'time now', ]:
        you(text)
        # print()
        if hour >= 12:
            printtext(hour, ":", minut, 'PM')
            speak(str(hour)+":"+str(minut)+'PM')
        else:
            printtext(hour, ":", minut, 'AM')
            speak(str(hour)+":"+str(minut)+'Am')

    elif "open gmail account" in text:
        rm = openpyxl.load_workbook("Chitkara_email.xlsx")

        sh1 = rm["Sheet1"]

        length = sh1.max_row

        # for read
        li_name = []
        for i in range(1, length+1):
            li_name.append(sh1.cell(i, 1).value)
        

        speak("which person you want to send email")
        name = command()
        print(name)
        speak("tell me subject")
        subject = command()
        print(subject)
        speak("what do you want to send")
        contant = command()
        
        you("which person you want to \nsend email\n->"+name+"\n"+"tell me subject\n->"+subject+"\n"+"what do you want to \nsend\n->"+contant)

        for i in range(1, length):
            ans = lower(li_name[i])
            if name == ans:
                email = sh1.cell(i+1, 2).value

        sendmail(email, subject, contant, name)

    elif ("weather of") in text or "weather in" in text:
        
        d = text.split()
       
        new_list = [d[i]+" " for i in range(2, len(d))]
        strrequest = ""
        for i in new_list:
            strrequest += i

        city_name =strrequest
        you(strrequest)
        data = requests.get("https://api.openweathermap.org/data/2.5/weather?q=" +
                            city_name+"&appid=882bef32516f8990d82b762cae54604c").json()

        weather = "weather is, "+data["weather"][0]["main"]
        Temp = "temperature is, " + \
            str(int(data["main"]["temp"]-273.5))+" degree Celsius"
        description = 'description is, '+data["weather"][0]['description']
        name = "Name, "+data["name"]+","
        code = "Code is, "+str(data["cod"])
        speak(weather+Temp+description+name+code)
        # you(text)
        printtext(weather+"\n"+Temp+"\n"+description+"\n"+name+"\n"+code)

    elif "temperature in" in text:
        # print(text)
        
        d = text.split()
       
        new_list = [d[i]+" " for i in range(2, len(d))]
        strrequest = ""
        for i in new_list:
            strrequest += i

        city_name =strrequest
        data = requests.get("https://api.openweathermap.org/data/2.5/weather?q=" +
                            city_name+"&appid=882bef32516f8990d82b762cae54604c").json()

        weather = "weather is, "+data["weather"][0]["main"]
        Temp = "temperature is, " + \
            str(int(data["main"]["temp"]-273.5))+" degree Celsius"
        description = 'description is, '+data["weather"][0]['description']
        name = "Name, "+data["name"]+","
        code = "Code is, "+str(data["cod"])
        
        speak(Temp+name)
        printtext(Temp+"\n"+name)
        

   
    elif text in ["brightness down"]:
        pct.set_brightness(30)

    elif text in ['open google', 'google open']:
        webbrowser.open('www.google.com')
        speak('opening google')

    elif text in ['open youtube', 'youtube open', "youtube"]:
        webbrowser.open('www.youtube.com')
        speak('opening youtube')

    elif text in ['open telegram', 'telegram open']:
        speak('opening telegram')
        webbrowser.open('www.telegram.com')

    elif text in ["open resource monitor", "resource monitor open"]:
        speak("opening resource monitor")
        subprocess.run('''resmon''')
    elif text in ["who made you","tum ko kisne banaya","who was created you"]:
        speak("i am made by lalit max")
        printtext("i am made by lalit max!")
    elif text in ["open computer management console", "computer management console open", "open computer management", "computer management open"]:
        speak("opening computer management console")
        subprocess.run('''compmgmtlauncher''')

    elif text in ['open whatsapp', 'whatsapp open']:
        speak('opening whatsapp')
        webbrowser.open('www.whatsap.com')

    elif text in ['open gmail', 'gmail open']:
        speak('opening gmail')
        webbrowser.open('www.gmail.com')

    elif text in ['open github', 'github open']:
        speak('opening github')
        webbrowser.open('www.github.com')

    elif text in ['open linkedin', 'linkedin open']:
        speak('opening linkedin')
        webbrowser.open('www.linkedin.com')

    elif text in ['open instagram', 'instagram open']:
        speak('opening insta gram')
        webbrowser.open('www.instagram.com')

    elif text in ['open music', 'play music', 'music open', 'music play']:
        os.startfile("wmplayer")

    elif text in ["open documents", "documents open"]:
        speak("opening documents")
        os.startfile("Documents")
    

    elif text in ["open wordpad", "wordpad open"]:
        speak("opening wordpad")
        os.startfile("wordpad")
   
   

    elif text in ["open ms paint", "ms paint open", "open paint", "paint open", "open microsoft paint"]:
        
        speak("opening ms paint")
        os.startfile("mspaint")
        printtext("Opened mspaint")

    elif text in ["open calculator", "calculator open"]:
        
        speak("opening calculator")
        os.startfile("calc")
        printtext("Opened calculator")
    elif text in ["open recycle bin","recycle bin"]:
        speak("opening recycle bin")
        os.startfile("shell:recyclebinfolder")
        printtext("Opened recycle bin")

    elif text in ["open task manager", "task manager open"]:
        # printtext(text)
        speak("opening task manager")
        os.startfile("taskmgr")
    elif text in ["open system information","system information","system information open","show system information","system information show"]:
        # printtext(text)
        speak("showing system information")
        os.startfile("MSINFO32")

    elif text in ["open notepad", "notepad open"]:
        # printtext(text)
        speak("opening notepad")
        os.startfile("notepad")

    elif text in ["open c drive", "open cdrive", "cdrive open", "c drive open"]:
        # printtext(text)
        speak("opening c drive")
        os.startfile("c:")
    elif text in ["open d drive", "open ddrive", "ddrive open", "d drive open"]:
        # printtext(text)
        speak("opening d drive")
        os.startfile("d:")
    elif text in ["open settings", "settings open", "open setting", "setting open"]:
        # printtext(text)
        speak("opening settings")
        os.startfile("ms-settings:")

    elif text in ["open microsoft store", "microsoft store open"]:
        # printtext(text)
        speak("opening microsoft store")
        os.startfile("ms-windows-store:")
    

    elif text in ["check windows version", "windows version check", "check window version", "what is window version"]:
        speak("checking your windows version")
        os.startfile("winver")

    elif text in ["open phone dialler", "phone dialler open"]:
        # printtext(text)
        speak("opening phone dialer")
        os.startfile("dialer")

    elif text in ["opens the temporary files folder", "open the temporary files folder", "open temporary files folder", "temporary files folder open", "open temporary files", "temporary file open", "open temporary file", "temporary files open"]:
        printtext(text)
        speak("opening the temporary files folder")
        os.startfile('''temp''')

    elif text in ["open camera", "camera open"]:
        printtext(text)
        speak("opening camera")
        os.startfile("microsoft.windows.camera:")

    elif text in ["Open the registry editor", "Opens the registry editor", "open registry editor", "resistry editor open"]:
        printtext(text)
        speak("Opening the Registry Editor")
        os.startfile("regedit")

    elif text in ["open disk management", "disk management open"]:
        printtext(text)
        speak("opening disk management")
        os.startfile("diskmgmt.msc")

    elif text in ["open this pc", "this pc open"]:
        printtext(text)
        speak("opening this pc")
        os.startfile('\"')

    elif text in ["open user account", "open user accounts", "user account open", "user accounts open"]:
        printtext(text)
        speak("opening user accounts5")
        os.startfile("netplwiz")

    elif text in ["open control Panel", "control Panel open"]:
        printtext(text)
        speak("opening control Panel")
        os.startfile("control")

    elif text in ["open device manager", "device manager open"]:
        printtext(text)
        speak("opening device manager")
        os.startfile('''devmgmt.msc''')

    elif text in ["open power option", "power option open"]:
        printtext(text)
        speak("opening power option")
        os.startfile('''powercfg.cpl''')

    elif text in ["open the directx diagnostic tool", "directx diagnostic tool open", "open directx diagnostic tool", "directx diagnostic tool open", "open diagnostic tool", "diagnostic tool open"]:
        printtext(text)
        subprocess.run('''dxdiag''')
        speak("opening the directx diagnostic tool")

    elif text in ["open programs and features", "programs and features open", "open uninstall setting", "uninstall setting open"]:
        printtext(text)
        os.startfile('''appwiz.cpl''')
        speak("opening programs and features")

    elif text in ["open character map", "character map open"]:
        printtext(text)
        os.startfile('charmap')
        speak("opening character map")

    elif text in ["open network connections", "network connections open", "open network connection", 'network connection open']:
        printtext(text)
        speak("opening network connections")
        os.startfile("ncpa.cpl")
    elif "on google" in text:
        printtext(text)
        li = text.split()
        st = ""
        for i in range(len(li)-2):
            st = st+li[i]+" "

        pwt.search(st)

    elif text in ["open keyboard", "keyboard open"]:
        printtext(text)
        os.startfile("osk")

    elif text in ["open mouse properties", "mouse properties open"]:
        printtext(text)
        speak("opening mouse properties")
        os.startfile("main.cpl")

    elif text in ["open remote desktop connection", "remote desktop connection open"]:
        printtext(text)
        speak("opening remote desktop connection")
        os.startfile("mstsc")

    elif text in ["open file explorer", "file explorer open"]:
        printtext(text)
        os.startfile("explorer")
        speak("opening file explorer")
    elif text in ["kaise ho"]:
        speak("i am fine, and how are you")
        printtext("i am fine, and how are you")

    elif text in ["shutdown", "shut down", "shutdown laptop", "laptop shut down", "laptop shutdown", "shut down laptop", "shutdown computer", "shut down computer", "computer shutdown", "computer shut down"]:
        printtext(text)
        a = subprocess.run("shutdown /s")
        speak("hello user i am going to shuting down your computer, wait some seconds. please don't touch any key until i shutdown computer")

    elif text in ["restart", "restart laptop", "laptop restart", "computer restart", "restart computer"]:
        printtext(text)
        a = subprocess.run("shutdown /r")
        speak("hello user i am going to shuting down your computer, wait some seconds. please don't touch any key until i restart computer")

    elif text in ["open powershell", "powershell open", "open power sale"]:
        printtext(text)
        os.startfile("powershell")
        speak("opening powershell")

    elif text in ["open excel", "excel open"]:
        printtext(text)
        speak("opening exel")
        os.startfile("excel")

    elif text in ["open chrome", "chrome open", "open google chrome", "google chrome open"]:
        printtext(text)
        speak("opening chrome")
        os.startfile("chrome")

    elif text in ["open brave", "brave open"]:
        printtext(text)
        os.startfile('Brave')
        speak("opening brave")

    elif text in ["open firefox", "firefox open"]:
        printtext(text)
        speak("opening firefox")
        os.startfile('firefox')

    elif text in ["open cmd", "cmd open", "open command prompt", "command prompt open"]:
        os.startfile("cmd")

   

    elif text in ["wishme", "wish me"]:
        print(text)
        wishme(text)
    elif text in ["open microsoft edge", "Microsoft edge open"]:
        printtext(text)
        speak("opening microsoft edge")
        os.startfile("msedge")

    elif text in ["open photoshop", "photoshop open"]:
        printtext(text)
        speak("opening photoshop")
        os.startfile("photoshop")

    elif text in ["open outlook", "outlook open"]:
        printtext(text)
        speak("opening outlook")
        os.startfile("outlook")
    elif text in ["open powerpoint", "powerpoint open"]:
        printtext(text)
        speak("opening powerpoint")
        os.startfile("powerpnt")

    elif 'how are you' in text:
        speak('i am fine, tell me how may i help you')
        printtext('i am fine.\ntell me how may i help you')
        
        
    elif 'what is your name' in text or "what's your name" in text:
        speak('''well, my name is maxy, i wish that everyone had a nickname as cool as mine, so plz keep small your name  ''')
        printtext('''well, my name's maxy" \ni wish that everyone\nhad a nickname as cool as mine\nso plz keep small and sort your name  ''')
        
    elif text in ['are you marry me', "will you marry me"]:
        speak("this is one of things, we'd both have to agree on i'd prefer to keep  our friendship as it is. ")
        printtext("this is one of things \nwe'd both have to agree\non i'd prefer to keep \nour friendship as it is.")
        
    elif text in ['what can you do for me']:
        speak("i can do all the work, which is in my might")
        printtext("i can do all the work \n which is in my might")
        
    elif text in ["do something for me"]:
        speak("Ask me any problem, i will try to solve it for you")
        printtext("Ask me any problem \ni will try to solve it \nfor you")
        
    elif text in ['date', "what's date", "what is date", "date", "what's the date today", "today date", "today's date", "what is the date", "what's the date"]:
    
        printtext(date)
        
        speak(date)
        
    elif text in ["tell me some jokes", "tell some jokes", "tell me some joke", "kucch joke sunao", "kuchh jokes sunao", 'tell me joke ', 'tell me jokes']:
        speak("Air hostess asked lalu Prasad yadav. Sir are you vegetarian or Non vegetarian, Lalu said I am indian. Air hostess said okay, Are you shakahari or mansahari, Lalu said hat sasuri I am Bihari")
        printtext("Air hostess asked lalu \nPrasad yadav. \nSir are you vegetarian or \nNon vegetarian \nLalu said I am indian \nAir hostess said okay, \nAre you shakahari or mansahari \nLalu said hat sasuri I am Bihari")
        
    elif "wikipedia" in text:
        result = wikipedia.summary(text, sentences=1)
        you(text)
        printtext(result)
        speak(result)
    elif   text in ["man nahi lag raha hai","how i feel good","not liking it"]:
        nm = random(1,10)
        speak("Ok so we think you will have to listen to some song then you will feel like it, well let's play the song for this")
        try:
            pwt.playonyt("Romantic song video injoy "+str(nm))
            speak('playing')
        except:
            speak("network Error Occurred ")
        printtext("Enjoy...")
    elif "print table of" in text or "table of" in text:

        nu = text.split()
        nu = int(nu[-1])
        strjee = ""
        for i in range(1, 11):
            strjee +=str(i*nu)+" "
        speak("The table of "+str(nu))
        printtext("The table of "+str(nu)+"\n\n"+strjee)

    elif "song of" in text:
        try:
            pwt.playonyt(text)
            speak('playing')
        except:
            speak("network Error Occurred ")
    elif "ka video" in text:
        try:
            pwt.playonyt(text)
            speak('playing')
        except:
            speak("network Error Occurred ")

    elif ("on youtube" in text) or ("on yt" in text):
        try:
            pwt.playonyt(text)
            speak('playing')
        except:
            speak("network Error Occurred ")
    elif "play video" in text:
        try:
            pwt.playonyt(text)
            speak('playing')
        except:
            speak("network Error Occurred ")

    elif "sum of" in text or "add" in text:
        
      
        text = text.split()
        li = []
        sum = 0
        for i in text:
            if i[0] >= '0' and i[0] <= '9':
                sum += float(i)
        speak("the answer is "+str(sum))
        printtext("the answer is", sum)
        
    elif ("even number " in text) or ("odd number" in text):
       
        text = text.split()
        li = []
        num = 0
        for i in text:
            if i[0] >= '0' and i[0] <= '9':
                num = int(i)
        if(num % 2 == 0):
            speak(str(num)+" is a even number")
            printtext(str(num)+" is a even number")
            
        else:
            speak(str(num)+" is a odd number")
            printtext(str(num)+" is a odd number")
            

    elif "area of circle" in text:
        text = text.split()
        li = []
        rad = 0
        for i in text:
            if i[0] >= '0' and i[0] <= '9':
                rad = float(i)
        area = 3.14*rad*rad
        speak("The area of circle is "+str(area))
        printtext("The area of circle is ", area)
        

    elif ("multiply" in text) or ("multiple" in text):
       
        print()
        text = text.split()
        multp = 1
        for i in text:
            if i[0] >= '0' and i[0] <= '9':
                multp *= float(i)
        speak("the answer is "+str(multp))
        print("the answer is", multp)
        
    elif ("divided" in text) or ("/" in text):
        
        text = text.split()
        li = []
        for i in text:
            if i[0] >= '0' and i[0] <= '9':
                li.append(float(i))
        a = li[0]
        b = li[1]
        printtext("the answer is"+str( a/b))
        speak("the answer is"+str(a/b))

    elif "how to make" in text:
        try:
            pwt.playonyt(text)
            speak("playing")
        except:
            speak("network Error Occurred ")

    elif text in ["do you know chitkara university"]:
        
        speak(  "yes i know chitkara university, it is best private university in the punjab ")
        print("yes i know chitkara university, \nit is the  best private university in the punjab ")
       
    elif "factorial" in text:
    
        fact = str(text)
        fact = fact.split()
        fact = int(fact[-1])

        fact = math.factorial(fact)
        speak("The answer is "+str(fact))
        printtext("The answer is "+str(fact))
        
    elif text in ["open coding ninjas", "coding ninjas open", "coding ninjas"]:
        webbrowser.open('https://www.codingninjas.com')
        printtext("opened coding ninjas")

    elif text in ["open vs code", "open visual studio code", "vs code open", "visual studio code open"]:
       
        speak("opening vs code editor")
        os.startfile("code")
        printtext("opened vs code")
    elif text in ["open photopea", "photopea open"]:
        
        speak("opening photopea")
        webbrowser.open("https://www.photopea.com")
        printtext("opened photopea")


    else:
        printtext("sorry i don't understand")
        speak("sorry i don't understand")

def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    base_path = getattr(sys, '_MEIPASS', os.path.dirname(
        os.path.abspath(__file__)))
    return os.path.join(base_path, relative_path)

    

if __name__ == '__main__':
    
    def karo():
            
            
            speak("Listening start")
        # while(True):
            text = command()
            if text in ["exit", "quit", "stop"]:
                exit()
            else:
                
                doooo(text)

    def func():
        x = var.get().lower()
        doooo(x)


    # for image not error when you convert py to exe
    

    # for title
    wind.title("Maxy")

#  for icon
    path1 = resource_path("mic.png")
    imj = PhotoImage(file=path1)
    wind.iconphoto(False, imj)

    # // window size

    wind.maxsize(width=1480, height=990)
    wind.minsize(width=1480, height=990)
    # wind.geometry("2000x2000")
    wind.colormapwindows()
   
    wind.config(background='#108cff')
    # label 1
    mylabel = Label(wind,font=("arial",20),bg='#108cff',fg='white',justify="left" ,border=1)
    mylabel.place(x=30,y=125)
    # show image
    path2 = resource_path("mic.png")
    my_pic5 = Image.open(path2)

    resize_pic5 = my_pic5.resize((100,102
                                  ), Image.ANTIALIAS)
    new_pic5 = ImageTk.PhotoImage(resize_pic5)
    
    lbl=Label(wind,bg="#108cff").place(x=290,y=60)
    
    # label 2
    mylabel2 = Label(wind,font=("arial",16 ),bg="#108cff",justify="left"   )
    mylabel2.place(x=1000,y=85)



    path3 = resource_path("searchicon.png")
    my_pic = Image.open(path3)


    resize_pic = my_pic.resize((53, 54), Image.ANTIALIAS)
    new_pic = ImageTk.PhotoImage(resize_pic)

    
    # wlcome to maxy
    path4 = resource_path("wctm.png")

    my_pic_wctm = Image.open(path4)

    resize_pic_wctm = my_pic_wctm.resize((388, 124), Image.ANTIALIAS)
    my_pic_wctm = ImageTk.PhotoImage(resize_pic_wctm)
    mylabel2wctm = Label(wind, image=my_pic_wctm,bg="#108cff" , )
    mylabel2wctm.pack(pady=0)

    def about():
        pass
    # menu bar
    menuBar= Menu(wind)
    wind.config(menu=menuBar)

    file_manu = Menu(menuBar)
    menuBar.add_cascade(label="About",menu=file_manu,activeforeground="#108cff",activebackground="#108cff")
    file_manu.add_cascade(label="Made by \n Lalit Max",command=about)
    


    var = StringVar()
    ent = Entry(wind, width=40, font=("Arial", 23), textvariable=var)
    ent.place(x=185, y=780)

    btn = Button(wind, image=new_pic, command=func,activebackground="#108cff",border=2,bg="white")
    btn.place(x=1219 ,y=780)

    btn2 = Button(wind, image=new_pic5, command=karo,activebackground="#108cff",bg="#108cff",border=2).place(x=690,y=610)
    
    wind.mainloop()
